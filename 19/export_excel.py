from datetime import datetime
from stock_data import fetch_all_stocks
from radar import LimitUpRadar
import pandas as pd
import os

EXCEL_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'reports')
os.makedirs(EXCEL_DIR, exist_ok=True)

def format_number(num, decimals=2):
    if pd.isna(num):
        return '-'
    return round(num, decimals)

def format_volume(volume):
    if pd.isna(volume):
        return '-'
    if volume >= 100000000:
        return f"{volume / 100000000:.2f}亿"
    elif volume >= 10000:
        return f"{volume / 10000:.2f}万"
    return f"{volume:.0f}"

def format_turnover(turnover):
    if pd.isna(turnover):
        return '-'
    if turnover >= 100000000:
        return f"{turnover / 100000000:.2f}亿"
    elif turnover >= 10000:
        return f"{turnover / 10000:.2f}万"
    return f"{turnover:.0f}"

def export_to_excel(analysis_results, filename=None):
    if not analysis_results:
        print("没有数据可导出")
        return None
    
    if filename is None:
        filename = f"涨停股票报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    filepath = os.path.join(EXCEL_DIR, filename)
    
    df = pd.DataFrame(analysis_results)
    
    export_df = pd.DataFrame({
        '序号': range(1, len(df) + 1),
        '股票代码': df['code'],
        '股票名称': df['name'],
        '当前价格(元)': df['price'].apply(lambda x: format_number(x, 2)),
        '涨跌额(元)': df['change_amount'].apply(lambda x: format_number(x, 2)),
        '涨幅(%)': df['change_pct'].apply(lambda x: format_number(x, 2)),
        '开盘价(元)': df['open'].apply(lambda x: format_number(x, 2)),
        '最高价(元)': df['high'].apply(lambda x: format_number(x, 2)),
        '最低价(元)': df['low'].apply(lambda x: format_number(x, 2)),
        '成交量': df['volume'].apply(format_volume),
        '成交额': df['turnover'].apply(format_turnover),
        '换手率(%)': df['turnover_rate'].apply(lambda x: format_number(x, 2)),
        '市盈率(PE)': df['pe'].apply(lambda x: format_number(x, 2)),
        '市净率(PB)': df['pb'].apply(lambda x: format_number(x, 2)),
        '涨停价(元)': df['high_limit'].apply(lambda x: format_number(x, 2)),
        '昨日收盘价(元)': df['prev_close'].apply(lambda x: format_number(x, 2)),
        '风险等级': df['risk_level'],
        '分析要点': df['analysis'].apply(lambda x: ', '.join(x) if isinstance(x, list) else str(x))
    })
    
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        export_df.to_excel(writer, sheet_name='涨停股票', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['涨停股票']
        
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        for row in worksheet.iter_rows(min_row=2, max_row=len(export_df) + 1):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for i, col in enumerate(export_df.columns, 1):
            col_letter = get_column_letter(i)
            max_length = max(
                len(str(export_df[col].iloc[j]) if pd.notna(export_df[col].iloc[j]) else '')
                for j in range(len(export_df))
            )
            if col == '股票名称':
                max_length = max(max_length, 12)
            elif col == '分析要点':
                max_length = max(max_length, 30)
            
            adjusted_width = min(max_length + 2, 40)
            worksheet.column_dimensions[col_letter].width = adjusted_width
        
        risk_fill_high = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        risk_fill_medium = PatternFill(start_color='FFD93D', end_color='FFD93D', fill_type='solid')
        risk_fill_low = PatternFill(start_color='6BCB77', end_color='6BCB77', fill_type='solid')
        
        risk_col = 17
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(export_df) + 1), start=2):
            risk_cell = row[risk_col - 1]
            risk_value = risk_cell.value
            if '高风险' in str(risk_value):
                risk_cell.fill = risk_fill_high
                risk_cell.font = Font(bold=True, color='FFFFFF')
            elif '中风险' in str(risk_value):
                risk_cell.fill = risk_fill_medium
                risk_cell.font = Font(bold=True)
            elif '低风险' in str(risk_value):
                risk_cell.fill = risk_fill_low
                risk_cell.font = Font(bold=True, color='FFFFFF')
        
        worksheet.row_dimensions[1].height = 25
        
        summary_df = pd.DataFrame([{
            '统计项目': '涨停股票总数',
            '数值': len(df),
        }, {
            '统计项目': '高风险股票',
            '数值': len(df[df['risk_level'] == '高风险']),
        }, {
            '统计项目': '中风险股票',
            '数值': len(df[df['risk_level'] == '中风险']),
        }, {
            '统计项目': '低风险股票',
            '数值': len(df[df['risk_level'] == '低风险']),
        }, {
            '统计项目': '平均涨幅',
            '数值': f"{df['change_pct'].mean():.2f}%",
        }, {
            '统计项目': '最高涨幅',
            '数值': f"{df['change_pct'].max():.2f}%",
        }, {
            '统计项目': '平均换手率',
            '数值': f"{df['turnover_rate'].mean():.2f}%",
        }, {
            '统计项目': '最高换手率',
            '数值': f"{df['turnover_rate'].max():.2f}%",
        }, {
            '统计项目': '扫描时间',
            '数值': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }])
        
        summary_df.to_excel(writer, sheet_name='统计摘要', index=False)
        summary_worksheet = writer.sheets['统计摘要']
        
        for cell in summary_worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        for row in summary_worksheet.iter_rows(min_row=2, max_row=len(summary_df) + 1):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        summary_worksheet.column_dimensions['A'].width = 20
        summary_worksheet.column_dimensions['B'].width = 20
    
    print(f"\nExcel报告已生成: {filepath}")
    return filepath

def main():
    print("涨停股票Excel报告导出工具")
    print("=" * 50)
    
    try:
        print("\n正在获取股票数据...")
        stocks = fetch_all_stocks()
        
        if stocks.empty:
            print("获取股票数据失败")
            return
        
        print(f"成功获取 {len(stocks)} 只股票数据")
        
        radar = LimitUpRadar()
        limit_up = radar.detect_limit_up(stocks)
        analysis = radar.analyze_limit_up_stocks(limit_up)
        
        print(f"检测到 {len(analysis)} 只涨停股票")
        
        if analysis:
            print("\n正在生成Excel报告...")
            filepath = export_to_excel(analysis)
            
            if filepath:
                print("\n" + "=" * 50)
                print("报告生成成功！")
                print(f"文件位置: {filepath}")
                print("=" * 50)
        else:
            print("没有涨停股票数据")
            
    except Exception as e:
        print(f"导出过程中出错: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()
